"""XLSX report generator using openpyxl.

Generates a multi-sheet workbook with:
- Summary sheet
- Component Scores sheet
- Theme Analysis sheet
- Per-component detail sheets
- All Remarks appendix
"""

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.charts import COMPONENT_ORDER, COMPONENT_SHORT_NAMES

logger = logging.getLogger("sehra.report_xlsx")

# Styling constants
TEAL_HEX = "0D7377"
RED_HEX = "CC3333"
HEADER_FILL = PatternFill(start_color=TEAL_HEX, end_color=TEAL_HEX, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TEAL_FONT = Font(bold=True, color=TEAL_HEX, size=11)
RED_FONT = Font(bold=True, color=RED_HEX, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COMPONENT_DISPLAY_NAMES = {
    "context": "Context",
    "policy": "Policy & Strategy",
    "service_delivery": "Service Delivery",
    "human_resources": "Human Resources",
    "supply_chain": "Supply Chain",
    "barriers": "Barriers",
}


def _style_header_row(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_column_widths(ws, min_width: int = 10, max_width: int = 60):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


def generate_xlsx_report(
    component_analyses: list[dict],
    header_info: dict,
    executive_summary: str = "",
    recommendations: str = "",
) -> io.BytesIO:
    """Generate a multi-sheet XLSX report.

    Args:
        component_analyses: List of component analysis dicts from DB
        header_info: {country, district, assessment_date}
        executive_summary: Optional executive summary text
        recommendations: Optional recommendations text

    Returns:
        BytesIO containing the XLSX file
    """
    logger.info("Generating XLSX report for %s", header_info.get("country", ""))
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    total_enablers = sum(a.get("enabler_count", 0) for a in component_analyses)
    total_barriers = sum(a.get("barrier_count", 0) for a in component_analyses)
    total_remarks = sum(len(a.get("qualitative_entries", [])) for a in component_analyses)
    grand_total = total_enablers + total_barriers
    readiness = round(total_enablers / grand_total * 100) if grand_total > 0 else 0

    summary_data = [
        ["SEHRA Analysis Report"],
        [""],
        ["Country", header_info.get("country", "")],
        ["District", header_info.get("district", "")],
        ["Assessment Date", header_info.get("assessment_date", "")],
        [""],
        ["Total Enablers", total_enablers],
        ["Total Barriers", total_barriers],
        ["Overall Readiness", f"{readiness}%"],
        ["Total Classified Remarks", total_remarks],
        [""],
    ]

    for row_data in summary_data:
        ws_summary.append(row_data)

    # Style title
    ws_summary["A1"].font = Font(bold=True, size=16, color=TEAL_HEX)
    ws_summary.merge_cells("A1:B1")

    # Style metric labels
    for row in [7, 8, 9, 10]:
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        val_cell = ws_summary.cell(row=row, column=2)
        if row == 7:
            val_cell.font = TEAL_FONT
        elif row == 8:
            val_cell.font = RED_FONT

    # Executive Summary
    if executive_summary:
        ws_summary.append(["Executive Summary"])
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, size=12)
        for para in executive_summary.strip().split("\n\n"):
            if para.strip():
                ws_summary.append([para.strip()])

    _auto_column_widths(ws_summary)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 80

    # --- Sheet 2: Component Scores ---
    ws_scores = wb.create_sheet("Component Scores")
    headers = ["Component", "Enablers", "Barriers", "Total", "Readiness %"]
    ws_scores.append(headers)
    _style_header_row(ws_scores, 1, len(headers))

    comp_lookup = {ca["component"]: ca for ca in component_analyses}
    for comp_key in COMPONENT_ORDER:
        ca = comp_lookup.get(comp_key)
        if not ca:
            continue
        e = ca.get("enabler_count", 0)
        b = ca.get("barrier_count", 0)
        t = e + b
        r = round(e / t * 100) if t > 0 else 0
        ws_scores.append([
            COMPONENT_DISPLAY_NAMES.get(comp_key, comp_key),
            e, b, t, f"{r}%"
        ])

    # Totals row
    ws_scores.append(["TOTAL", total_enablers, total_barriers, grand_total, f"{readiness}%"])
    total_row = ws_scores.max_row
    for col in range(1, 6):
        ws_scores.cell(row=total_row, column=col).font = Font(bold=True)

    _auto_column_widths(ws_scores)
    ws_scores.freeze_panes = "A2"

    # --- Sheet 3: Theme Analysis ---
    ws_themes = wb.create_sheet("Theme Analysis")

    # Build theme × component pivot
    theme_data = {}
    for ca in component_analyses:
        comp = ca.get("component", "")
        for entry in ca.get("qualitative_entries", []):
            theme = entry.get("theme", "Other")
            if theme not in theme_data:
                theme_data[theme] = {}
            theme_data[theme][comp] = theme_data[theme].get(comp, 0) + 1

    comp_headers = [COMPONENT_DISPLAY_NAMES.get(c, c) for c in COMPONENT_ORDER]
    ws_themes.append(["Theme"] + comp_headers + ["Total"])
    _style_header_row(ws_themes, 1, len(COMPONENT_ORDER) + 2)

    for theme in sorted(theme_data.keys()):
        row = [theme]
        total = 0
        for comp_key in COMPONENT_ORDER:
            count = theme_data[theme].get(comp_key, 0)
            row.append(count)
            total += count
        row.append(total)
        ws_themes.append(row)

    _auto_column_widths(ws_themes)
    ws_themes.freeze_panes = "B2"

    # --- Per-Component Sheets ---
    for comp_key in COMPONENT_ORDER:
        ca = comp_lookup.get(comp_key)
        if not ca:
            continue

        sheet_name = COMPONENT_DISPLAY_NAMES.get(comp_key, comp_key)[:31]  # max sheet name length
        ws_comp = wb.create_sheet(sheet_name)

        # Header info
        ws_comp.append([sheet_name])
        ws_comp.cell(row=1, column=1).font = Font(bold=True, size=14, color=TEAL_HEX)
        ws_comp.append([f"Enablers: {ca.get('enabler_count', 0)}", f"Barriers: {ca.get('barrier_count', 0)}"])
        ws_comp.append([])

        # Entries table
        entries = ca.get("qualitative_entries", [])
        if entries:
            entry_headers = ["Item ID", "Theme", "Classification", "Confidence", "Remark"]
            ws_comp.append(entry_headers)
            _style_header_row(ws_comp, ws_comp.max_row, len(entry_headers))

            for entry in entries:
                ws_comp.append([
                    entry.get("item_id", ""),
                    entry.get("theme", ""),
                    entry.get("classification", ""),
                    f"{entry.get('confidence', 0):.0%}",
                    (entry.get("remark_text", "") or "")[:500],
                ])

            # Conditional formatting for classification
            for row_idx in range(5, ws_comp.max_row + 1):
                cls_cell = ws_comp.cell(row=row_idx, column=3)
                if cls_cell.value in ("enabler", "strength"):
                    cls_cell.font = Font(color=TEAL_HEX)
                elif cls_cell.value in ("barrier", "weakness"):
                    cls_cell.font = Font(color=RED_HEX)

        # Report sections
        sections = ca.get("report_sections", {})
        if sections:
            ws_comp.append([])
            for sec_type, sec_data in sections.items():
                content = sec_data.get("content", "")
                if content:
                    label = sec_type.replace("_", " ").title()
                    ws_comp.append([label])
                    ws_comp.cell(row=ws_comp.max_row, column=1).font = Font(bold=True)
                    ws_comp.append([content[:2000]])
                    ws_comp.append([])

        _auto_column_widths(ws_comp)

    # --- All Remarks Sheet ---
    ws_remarks = wb.create_sheet("All Remarks")
    remark_headers = ["Component", "Item ID", "Theme", "Classification", "Confidence", "Remark"]
    ws_remarks.append(remark_headers)
    _style_header_row(ws_remarks, 1, len(remark_headers))

    for comp_key in COMPONENT_ORDER:
        ca = comp_lookup.get(comp_key)
        if not ca:
            continue
        for entry in ca.get("qualitative_entries", []):
            ws_remarks.append([
                COMPONENT_DISPLAY_NAMES.get(comp_key, comp_key),
                entry.get("item_id", ""),
                entry.get("theme", ""),
                entry.get("classification", ""),
                f"{entry.get('confidence', 0):.0%}",
                (entry.get("remark_text", "") or "")[:500],
            ])

    _auto_column_widths(ws_remarks)
    ws_remarks.freeze_panes = "A2"

    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("XLSX report generated")
    return buf
